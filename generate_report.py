"""
Automatický Excel Report Generator
Vezme CSV data → vygeneruje přehledný .xlsx report s grafy
"""

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
import io
from datetime import datetime
import sys
import os


# ── Barvy ────────────────────────────────────────────────────────────────────
CLR = {
    "primary":    "1E3A5F",   # tmavě modrá
    "secondary":  "2E86AB",   # střední modrá
    "accent":     "F4A261",   # oranžová
    "success":    "2A9D8F",   # zelená
    "light_bg":   "EBF2FA",   # světle modrá plocha
    "header_bg":  "1E3A5F",   # header řádek
    "alt_row":    "F0F6FF",   # střídání řádků
    "white":      "FFFFFF",
    "border":     "B0C4DE",
    "text_dark":  "1A1A2E",
}

def clr(hex_str):
    return PatternFill("solid", start_color=hex_str, fgColor=hex_str)

def border(style="thin"):
    s = Side(style=style, color=CLR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, size=11, color="1A1A2E", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Načtení a zpracování dat ──────────────────────────────────────────────────
def load_and_process(csv_path: str) -> dict:
    df = pd.read_csv(csv_path, parse_dates=["Datum"])
    df["Zisk"]    = df["Tržby"] - df["Náklady"]
    df["Marže %"] = (df["Zisk"] / df["Tržby"] * 100).round(1)
    df["Měsíc"]   = df["Datum"].dt.to_period("M").astype(str)

    monthly = df.groupby("Měsíc").agg(
        Tržby=("Tržby", "sum"),
        Náklady=("Náklady", "sum"),
        Zisk=("Zisk", "sum"),
        Kusy=("Kusy", "sum"),
    ).reset_index()
    monthly["Marže %"] = (monthly["Zisk"] / monthly["Tržby"] * 100).round(1)

    by_cat = df.groupby("Kategorie").agg(
        Tržby=("Tržby", "sum"),
        Zisk=("Zisk", "sum"),
        Kusy=("Kusy", "sum"),
    ).reset_index()
    by_cat["Marže %"] = (by_cat["Zisk"] / by_cat["Tržby"] * 100).round(1)

    by_prod = df.groupby("Produkt").agg(
        Tržby=("Tržby", "sum"),
        Zisk=("Zisk", "sum"),
        Kusy=("Kusy", "sum"),
    ).reset_index().sort_values("Tržby", ascending=False)

    kpis = {
        "Celkové tržby":   df["Tržby"].sum(),
        "Celkový zisk":    df["Zisk"].sum(),
        "Průměrná marže":  df["Marže %"].mean(),
        "Prodané kusy":    df["Kusy"].sum(),
        "Počet produktů":  df["Produkt"].nunique(),
        "Sledované měsíce": df["Měsíc"].nunique(),
    }
    return dict(df=df, monthly=monthly, by_cat=by_cat, by_prod=by_prod, kpis=kpis)


# ── Matplotlib grafy jako PNG do paměti ──────────────────────────────────────
def make_charts(data: dict) -> dict:
    monthly = data["monthly"]
    by_cat  = data["by_cat"]
    by_prod = data["by_prod"]
    images  = {}

    COLORS = ["#1E3A5F", "#2E86AB", "#F4A261", "#2A9D8F", "#E76F51", "#457B9D"]

    # Graf 1 — Měsíční tržby vs náklady (skupinový sloupcový)
    fig, ax = plt.subplots(figsize=(9, 4.2))
    x = range(len(monthly))
    w = 0.35
    ax.bar([i - w/2 for i in x], monthly["Tržby"],  width=w, color="#2E86AB", label="Tržby",  zorder=3)
    ax.bar([i + w/2 for i in x], monthly["Náklady"], width=w, color="#F4A261", label="Náklady", zorder=3)
    ax.set_xticks(list(x)); ax.set_xticklabels(monthly["Měsíc"], rotation=30, ha="right", fontsize=9)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v/1000:.0f}k Kč"))
    ax.set_title("Měsíční tržby vs náklady", fontsize=13, fontweight="bold", pad=12)
    ax.legend(framealpha=0.9); ax.grid(axis="y", alpha=0.3, zorder=0)
    ax.set_facecolor("#F8FBFF"); fig.patch.set_facecolor("white")
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    images["monthly_bar"] = buf; plt.close(fig)

    # Graf 2 — Zisk v čase (liniový)
    fig, ax = plt.subplots(figsize=(9, 3.8))
    ax.plot(monthly["Měsíc"], monthly["Zisk"], marker="o", color="#1E3A5F",
            linewidth=2.5, markersize=7, zorder=3)
    ax.fill_between(range(len(monthly)), monthly["Zisk"], alpha=0.12, color="#1E3A5F")
    ax.set_xticks(range(len(monthly))); ax.set_xticklabels(monthly["Měsíc"], rotation=30, ha="right", fontsize=9)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v/1000:.0f}k Kč"))
    ax.set_title("Vývoj zisku", fontsize=13, fontweight="bold", pad=12)
    ax.grid(alpha=0.3); ax.set_facecolor("#F8FBFF"); fig.patch.set_facecolor("white")
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    images["profit_line"] = buf; plt.close(fig)

    # Graf 3 — Tržby podle kategorií (koláčový)
    fig, ax = plt.subplots(figsize=(5.5, 4.5))
    wedge_colors = COLORS[:len(by_cat)]
    wedges, texts, autotexts = ax.pie(
        by_cat["Tržby"], labels=by_cat["Kategorie"],
        autopct="%1.1f%%", colors=wedge_colors,
        startangle=140, pctdistance=0.75,
        wedgeprops=dict(edgecolor="white", linewidth=2)
    )
    for t in autotexts: t.set_fontsize(10); t.set_color("white"); t.set_fontweight("bold")
    ax.set_title("Tržby dle kategorie", fontsize=13, fontweight="bold")
    fig.patch.set_facecolor("white")
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    images["cat_pie"] = buf; plt.close(fig)

    # Graf 4 — Top produkty
    fig, ax = plt.subplots(figsize=(7.5, 4))
    colors_prod = [COLORS[i % len(COLORS)] for i in range(len(by_prod))]
    bars = ax.barh(by_prod["Produkt"], by_prod["Tržby"], color=colors_prod, zorder=3)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v/1000:.0f}k"))
    ax.set_xlabel("Tržby (Kč)", fontsize=9)
    ax.set_title("Tržby podle produktů", fontsize=13, fontweight="bold", pad=12)
    ax.grid(axis="x", alpha=0.3, zorder=0); ax.set_facecolor("#F8FBFF")
    fig.patch.set_facecolor("white")
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    images["prod_bar"] = buf; plt.close(fig)

    return images


# ── Excel tvorba ──────────────────────────────────────────────────────────────
def write_cell(ws, row, col, value, bold=False, size=11, color="1A1A2E",
               bg=None, h_align="left", italic=False, number_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h_align, wrap=wrap)
    c.border    = border()
    if bg:   c.fill = clr(bg)
    if number_fmt: c.number_format = number_fmt
    return c


def add_section_title(ws, row, col_start, col_end, title):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.font      = font(bold=True, size=12, color=CLR["white"])
    c.fill      = clr(CLR["primary"])
    c.alignment = align(h="left")
    c.border    = border()


def build_excel(data: dict, images: dict, output_path: str):
    wb = Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard"
    ws_dash.sheet_view.showGridLines = False
    ws_dash.column_dimensions["A"].width = 2

    for col in range(2, 18):
        ws_dash.column_dimensions[get_column_letter(col)].width = 13

    # Záhlaví reportu
    ws_dash.row_dimensions[1].height = 8
    ws_dash.merge_cells("B2:Q4")
    c = ws_dash["B2"]
    c.value     = "📈  PRODEJNÍ REPORT — 2024"
    c.font      = Font(name="Arial", bold=True, size=20, color=CLR["white"])
    c.fill      = clr(CLR["primary"])
    c.alignment = align(h="center", v="center")

    ws_dash.merge_cells("B5:Q5")
    c = ws_dash["B5"]
    c.value     = f"Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   Zdroj: data.csv"
    c.font      = font(size=9, color=CLR["white"], italic=True)
    c.fill      = clr(CLR["secondary"])
    c.alignment = align(h="center")

    ws_dash.row_dimensions[6].height = 10

    # KPI boxy (řádky 7–10)
    kpis = data["kpis"]
    kpi_data = [
        ("💰 Celkové tržby",    f"{kpis['Celkové tržby']:,.0f} Kč",   CLR["primary"]),
        ("📦 Celkový zisk",     f"{kpis['Celkový zisk']:,.0f} Kč",    CLR["secondary"]),
        ("📊 Průměrná marže",   f"{kpis['Průměrná marže']:.1f} %",    CLR["success"]),
        ("🛒 Prodané kusy",     f"{int(kpis['Prodané kusy'])} ks",     CLR["accent"]),
    ]

    col_starts = [2, 6, 10, 14]
    for (label, value, bg), cs in zip(kpi_data, col_starts):
        ws_dash.merge_cells(start_row=7, start_column=cs, end_row=7, end_column=cs+3)
        ws_dash.merge_cells(start_row=8, start_column=cs, end_row=8, end_column=cs+3)
        ws_dash.merge_cells(start_row=9, start_column=cs, end_row=9, end_column=cs+3)

        cl = ws_dash.cell(row=7, column=cs, value=label)
        cl.font = font(size=10, color=CLR["white"]); cl.fill = clr(bg); cl.alignment = align(h="center")

        cv = ws_dash.cell(row=8, column=cs, value=value)
        cv.font = Font(name="Arial", bold=True, size=16, color=CLR["white"])
        cv.fill = clr(bg); cv.alignment = align(h="center", v="center")
        ws_dash.row_dimensions[8].height = 30

        ce = ws_dash.cell(row=9, column=cs, value="")
        ce.fill = clr(bg)

    ws_dash.row_dimensions[10].height = 10

    # Vložení grafů
    def insert_img(ws, buf, cell, width_cm, height_cm):
        img = XLImage(buf)
        img.width  = width_cm * 37.8
        img.height = height_cm * 37.8
        ws.add_image(img, cell)

    insert_img(ws_dash, images["monthly_bar"], "B11",  16, 7)
    insert_img(ws_dash, images["profit_line"], "J11",  16, 7)
    insert_img(ws_dash, images["cat_pie"],     "B29",  10, 8)
    insert_img(ws_dash, images["prod_bar"],    "J29",  13, 8)

    # ── Sheet 2: Měsíční přehled ──────────────────────────────────────────────
    ws_m = wb.create_sheet("📅 Měsíce")
    ws_m.sheet_view.showGridLines = False
    ws_m.column_dimensions["A"].width = 2

    headers = ["Měsíc", "Tržby (Kč)", "Náklady (Kč)", "Zisk (Kč)", "Marže %", "Kusy"]
    widths  = [15, 18, 18, 18, 12, 10]
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    ws_m.row_dimensions[1].height = 8
    add_section_title(ws_m, 2, 2, 7, "  Měsíční přehled prodejů")

    for i, h in enumerate(headers, start=2):
        write_cell(ws_m, 3, i, h, bold=True, size=10, color=CLR["white"], bg=CLR["secondary"], h_align="center")

    monthly = data["monthly"]
    for r_idx, row in monthly.iterrows():
        excel_row = r_idx + 4
        bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
        vals = [row["Měsíc"], row["Tržby"], row["Náklady"], row["Zisk"], row["Marže %"]/100, int(row["Kusy"])]
        fmts = [None, '#,##0 "Kč"', '#,##0 "Kč"', '#,##0 "Kč"', '0.0%', None]
        for c_idx, (v, fmt) in enumerate(zip(vals, fmts), start=2):
            write_cell(ws_m, excel_row, c_idx, v, bg=bg, h_align="center", number_fmt=fmt)

    # Souhrnný řádek
    last = len(monthly) + 4
    write_cell(ws_m, last, 2, "CELKEM", bold=True, bg=CLR["light_bg"], h_align="center")
    for col_idx, col_letter in enumerate(["C","D","E"], start=3):
        ws_m.cell(row=last, column=col_idx).value = f"=SUM({get_column_letter(col_idx)}4:{get_column_letter(col_idx)}{last-1})"
        ws_m.cell(row=last, column=col_idx).font  = font(bold=True)
        ws_m.cell(row=last, column=col_idx).fill  = clr(CLR["light_bg"])
        ws_m.cell(row=last, column=col_idx).alignment = align(h="center")
        ws_m.cell(row=last, column=col_idx).border = border()
        if col_idx < 6:
            ws_m.cell(row=last, column=col_idx).number_format = '#,##0 "Kč"'
    write_cell(ws_m, last, 6, "", bg=CLR["light_bg"])
    write_cell(ws_m, last, 7, f"=SUM(G4:G{last-1})", bold=True, bg=CLR["light_bg"], h_align="center")

    # ── Sheet 3: Produkty ─────────────────────────────────────────────────────
    ws_p = wb.create_sheet("🏷️ Produkty")
    ws_p.sheet_view.showGridLines = False
    ws_p.column_dimensions["A"].width = 2

    p_headers = ["Produkt", "Kategorie", "Tržby (Kč)", "Zisk (Kč)", "Marže %", "Kusy"]
    p_widths  = [25, 16, 18, 18, 12, 10]
    for i, w in enumerate(p_widths, start=2):
        ws_p.column_dimensions[get_column_letter(i)].width = w

    ws_p.row_dimensions[1].height = 8
    add_section_title(ws_p, 2, 2, 7, "  Přehled podle produktů")

    for i, h in enumerate(p_headers, start=2):
        write_cell(ws_p, 3, i, h, bold=True, size=10, color=CLR["white"], bg=CLR["secondary"], h_align="center")

    df = data["df"]
    by_prod_full = df.groupby(["Produkt", "Kategorie"]).agg(
        Tržby=("Tržby","sum"), Zisk=("Zisk","sum"), Kusy=("Kusy","sum")
    ).reset_index().sort_values("Tržby", ascending=False)
    by_prod_full["Marže"] = by_prod_full["Zisk"] / by_prod_full["Tržby"]

    for r_idx, row in enumerate(by_prod_full.itertuples(index=False), start=4):
        bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
        vals = [row.Produkt, row.Kategorie, row.Tržby, row.Zisk, row.Marže, int(row.Kusy)]
        fmts = [None, None, '#,##0 "Kč"', '#,##0 "Kč"', '0.0%', None]
        for c_idx, (v, fmt) in enumerate(zip(vals, fmts), start=2):
            write_cell(ws_p, r_idx, c_idx, v, bg=bg, h_align="center" if c_idx > 3 else "left", number_fmt=fmt)

    # ── Sheet 4: Surová data ──────────────────────────────────────────────────
    ws_raw = wb.create_sheet("📋 Zdrojová data")
    ws_raw.sheet_view.showGridLines = False
    ws_raw.column_dimensions["A"].width = 2

    raw_headers = ["Datum", "Produkt", "Kategorie", "Tržby (Kč)", "Náklady (Kč)", "Zisk (Kč)", "Marže %", "Kusy"]
    raw_widths  = [14, 22, 15, 16, 16, 14, 12, 8]
    for i, w in enumerate(raw_widths, start=2):
        ws_raw.column_dimensions[get_column_letter(i)].width = w

    ws_raw.row_dimensions[1].height = 8
    add_section_title(ws_raw, 2, 2, 9, "  Zdrojová data")
    for i, h in enumerate(raw_headers, start=2):
        write_cell(ws_raw, 3, i, h, bold=True, size=10, color=CLR["white"], bg=CLR["secondary"], h_align="center")

    df_disp = data["df"].sort_values("Datum").reset_index(drop=True)
    for r_idx, (_, row) in enumerate(df_disp.iterrows(), start=4):
        bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
        datum = row["Datum"].strftime("%d.%m.%Y") if hasattr(row["Datum"], "strftime") else str(row["Datum"])
        marze = row["Zisk"] / row["Tržby"]
        vals  = [datum, row["Produkt"], row["Kategorie"], row["Tržby"], row["Náklady"], row["Zisk"], marze, int(row["Kusy"])]
        fmts  = [None, None, None, '#,##0 "Kč"', '#,##0 "Kč"', '#,##0 "Kč"', '0.0%', None]
        haligns = ["center","left","left","center","center","center","center","center"]
        for c_idx, (v, fmt, ha) in enumerate(zip(vals, fmts, haligns), start=2):
            write_cell(ws_raw, r_idx, c_idx, v, bg=bg, h_align=ha, number_fmt=fmt)

    wb.save(output_path)
    return output_path


# ── Hlavní funkce ─────────────────────────────────────────────────────────────
def generate_report(csv_path: str, output_path: str = "report.xlsx"):
    print(f"📂 Načítám data z: {csv_path}")
    data   = load_and_process(csv_path)
    print("📊 Generuji grafy...")
    images = make_charts(data)
    print("📝 Sestavuji Excel report...")
    build_excel(data, images, output_path)
    print(f"✅ Report uložen: {output_path}")
    kpis = data["kpis"]
    print(f"\n   Tržby celkem:  {kpis['Celkové tržby']:,.0f} Kč")
    print(f"   Zisk celkem:   {kpis['Celkový zisk']:,.0f} Kč")
    print(f"   Prům. marže:   {kpis['Průměrná marže']:.1f} %")
    print(f"   Prodáno kusů:  {int(kpis['Prodané kusy'])}")


if __name__ == "__main__":
    csv_in  = sys.argv[1] if len(sys.argv) > 1 else "data.csv"
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else "report.xlsx"
    generate_report(csv_in, xlsx_out)
